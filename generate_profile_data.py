import json
from pathlib import Path

import openpyxl

base = Path(r"C:\Users\dcollazos.SEGITTUR\Documents\ChatGPT\Dashboard PID Septiembre")
src = base / "datos" / "Simulacion_Respuestas_Encuestas_Enoturismo_Jerez.xlsx"
dest = base / "datos" / "sim_diagnostico_turista.js"

wb = openpyxl.load_workbook(src, data_only=True)
ws = wb["Sim. Diagnostico Turista"]
headers = [ws.cell(4, c).value for c in range(1, ws.max_column + 1)]
records = []
for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
    item = {}
    for h, v in zip(headers, row):
        if h:
            item[str(h)] = v
    records.append(item)

payload = {
    "sheet": "Sim. Diagnostico Turista",
    "headers": headers,
    "rows": records,
}

dest.write_text("window.SIM_DIAG_TURISTA = " + json.dumps(payload, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
